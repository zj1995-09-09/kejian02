
import openpyxl

# 打开excel文件

workbook = openpyxl.open('cases.xlsx')
print(workbook)

# 选择 sheet表格， worksheet
# 通过表格的名称去获取：类似于字典的操作
ws  = workbook['register']
print(ws)

# 获取数据，某个单元格的数据
cell = ws.cell(row=1, column=2)
print(cell.value)
print(ws.cell(row=3, column=1).value)

# 某一行
print(ws[2])

# 获取所有的行
print(list(ws.rows))

# 获取所有的数据
print(list(ws.values))

